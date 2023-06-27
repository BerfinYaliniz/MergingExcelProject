import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import calendar
import locale

# Türkçe ay adları için locale ayarlaması
locale.setlocale(locale.LC_TIME, 'tr_TR')

klasor_yolu = input("Excel dosyalarının bulunduğu ana klasör yolunu girin: ")

if klasor_yolu.startswith('"') and klasor_yolu.endswith('"'):
    klasor_yolu = klasor_yolu[1:-1]

tarih = datetime.now().strftime("%d-%m-%Y")
ay = calendar.month_name[int(datetime.now().strftime("%m"))].capitalize()
cikti_dosyasi = f"caka-{ay}-ürün-çıkış-listesi-{tarih}.xlsx"

veri_listesi = []

for root, dirs, files in os.walk(klasor_yolu):
    for dosya_adi in files:
        if dosya_adi.endswith('.xlsx') or dosya_adi.endswith('.xls'):
            dosya_yolu = os.path.join(root, dosya_adi)
            excel_veri = pd.read_excel(dosya_yolu)
            satir_sayisi = excel_veri.shape[0]
            klasor_adi = os.path.basename(root)

            # Klasördeki metin dosyasının adını alın
            metin_dosyasi_adi = ""
            for dosya in os.listdir(root):
                if dosya.endswith('.txt'):
                    metin_dosyasi_adi = dosya
                    break

            # Klasör adını metin dosyasının adıyla değiştirme
            klasor_adi = metin_dosyasi_adi.replace(".txt", "")

            # Dosya adını uzantıları kaldırarak güncelleme
            dosya_adi = dosya_adi.split(".")[0]

            veri_listesi.append((klasor_adi, dosya_adi, satir_sayisi))

for veri in veri_listesi:
    klasor_adi, dosya_adi, satir_sayisi = veri
    print(f"{klasor_adi} stok kodundaki {dosya_adi} ürün: {satir_sayisi} adet")

df = pd.DataFrame(veri_listesi, columns=['Stok Kodu', 'Ürün Adı', 'Ürün Adedi'])

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Veriler'

bold_font = Font(bold=True)

basliklar = df.columns.tolist()
for col_num, baslik in enumerate(basliklar, start=1):
    col_letter = get_column_letter(col_num)
    cell = worksheet[f"{col_letter}1"]
    cell.value = baslik
    cell.font = bold_font

for index, row in df.iterrows():
    for col_num, value in enumerate(row, start=1):
        col_letter = get_column_letter(col_num)
        if col_num == 2:
            value = str(value).split(".")[0]  # Uzantıyı kaldırma
        worksheet[f"{col_letter}{index + 2}"] = value

for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 8) * 1.2
    worksheet.column_dimensions[column].width = adjusted_width

workbook.save(cikti_dosyasi)

print(f"Çıktı başarıyla {cikti_dosyasi} dosyasına yazıldı.")
input("İşlem tamamlandı. Çıkmak için Enter tuşuna basın.")
